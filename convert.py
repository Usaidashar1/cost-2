"""
Azure Pricing Calculator → Cost Estimation Workbook (Production Ready)
"""
import re
import sys
import logging
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
log = logging.getLogger(__name__)

THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NUM = "#,##0.00"

def _f(bold=False, italic=False, size=11, color="000000"):
    return Font(name="Calibri", bold=bold, italic=italic, size=size, color=color)

def _fill(h): return PatternFill("solid", fgColor=h)
def _al(h="left", v="center", w=False): return Alignment(horizontal=h, vertical=v, wrap_text=w)

def hdr(c, v, wrap=False):
    c.value=v; c.font=Font(name="Calibri",bold=True,size=10,color="FFFFFF")
    c.fill=_fill("4472C4"); c.alignment=_al("center","center",wrap); c.border=BORDER

def dat(c, v, bold=False, italic=False, align="left", color="000000"):
    c.value=v; c.font=_f(bold,italic,color=color)
    c.alignment=_al(align,"center",True); c.border=BORDER
    if isinstance(v,(int,float)) and align=="right": c.number_format=NUM

def tot(c, v):
    c.value=v; c.font=_f(bold=True); c.fill=_fill("D9E1F2")
    c.alignment=_al("right","center"); c.border=BORDER
    if isinstance(v,(int,float)): c.number_format=NUM

def widths(ws, d):
    for col, w in d.items(): ws.column_dimensions[col].width = w

SVC_MAP = {
    "virtual machines": "Virtual Machines",
    "managed disks": "Managed Disks",
    "azure backup": "Azure Backup",
    "backup": "Azure Backup",
    "load balancer": "Load Balancer",
    "load balancers": "Load Balancer",
    "application gateway": "Application Gateway",
    "azure firewall": "Azure Firewall",
    "vpn gateway": "VPN Gateway",
    "storage": "Storage Accounts",
    "storage accounts": "Storage Accounts",
    "sql database": "SQL",
    "azure sql": "SQL",
    "ip addresses": "Public IP Addresses",
    "bandwidth": "Bandwidth",
    "azure monitor": "Azure Monitor",
    "key vault": "Key Vault",
}

SHEET_ORDER = [
    "Virtual Machines", "Managed Disks", "Public IP Addresses",
    "Load Balancer", "Application Gateway", "Azure Firewall", "VPN Gateway",
    "Storage Accounts", "Azure Backup", "SQL", "Bandwidth",
    "Azure Monitor", "Key Vault", "Others",
]

SKIP = {"support", "disclaimer", "total", "licensing program", "billing account", "billing profile"}

def arm_region(display):
    val = str(display).lower().strip()
    return val.replace(" ", "")

# ── API Setup & Helpers ──────────────────────────────────────────────────
API = "https://prices.azure.com/api/retail/prices"

def get_http_session():
    session = requests.Session()
    retries = Retry(total=5, backoff_factor=1, status_forcelist=[429, 500, 502, 503, 504])
    session.mount("https://", HTTPAdapter(max_retries=retries))
    return session

def _api(session, cache, filt, currency="INR"):
    key = filt + currency
    if key in cache: return cache[key]
    try:
        r = session.get(API, params={"api-version":"2023-01-01-preview","$filter":filt,"currencyCode":currency}, timeout=15)
        r.raise_for_status()
        items = r.json().get("Items", [])
        cache[key] = items
        return items
    except Exception as e:
        log.warning(f"API Fetch Error: {e}")
        return []

def _hourly_to_monthly(price): return price * 730

def detect_os(desc):
    desc_l = desc.lower()
    if "hybrid benefit" in desc_l: return "Linux" 
    if "red hat" in desc_l or "rhel" in desc_l: return "Red Hat"
    if "suse" in desc_l or "sles" in desc_l: return "SUSE"
    if "linux" in desc_l or "ubuntu" in desc_l or "centos" in desc_l: return "Linux"
    if "windows" in desc_l: return "Windows"
    return "Windows" 

def detect_sql_license(desc):
    desc_l = desc.lower()
    if "sql enterprise" in desc_l: return "SQL Enterprise License"
    if "sql standard"   in desc_l: return "SQL Standard License"
    if "sql web"        in desc_l: return "SQL Web License"
    if "sql developer"  in desc_l: return "SQL Developer License"
    if "sql" in desc_l: return "SQL License"
    return None

def get_exact_license_name(desc, os_type):
    parts = re.split(r'[,;]', desc)
    sql_name = os_name = None
    for p in parts:
        p_lower = p.lower()
        if "sql" in p_lower:
            sql_name = re.sub(r'\s*\([^)]*\)', '', p).strip() 
        if os_type in ["Red Hat", "SUSE"] and ("red hat" in p_lower or "rhel" in p_lower or "suse" in p_lower or "sles" in p_lower):
            os_name = re.sub(r'\s*\([^)]*\)', '', p).strip()
    return sql_name, os_name

def get_vm_pricing(session, cache, sku, region_display, is_spot, currency="INR"):
    region = arm_region(region_display)
    result = {"compute_payg": None, "windows_license": None, "compute_ri1": None, "compute_ri3": None}

    try:
        def fetch(s, ptype): 
            return _api(session, cache, f"armSkuName eq '{s}' and armRegionName eq '{region}' and priceType eq '{ptype}'", currency)
        
        def get_payg(s):
            items = fetch(s, "Consumption")
            if items: return items
            items = fetch(s.replace("s_v", "_v").replace("s_V", "_V"), "Consumption")
            if items: return items
            items = fetch(s.replace("ds_v", "d_v").replace("ds_V", "d_V"), "Consumption")
            return items or []

        def get_ri(s):
            items = fetch(s, "Reservation")
            if items: return items
            items = fetch(s.replace("s_v", "_v").replace("s_V", "_V"), "Reservation")
            if items: return items
            items = fetch(s.replace("ds_v", "d_v").replace("ds_V", "d_V"), "Reservation")
            return items or []

        payg_items = get_payg(sku)
        ri_items = get_ri(sku)
        
        if not payg_items:
            base_sku = re.sub(r'-\d+', '', sku)
            if base_sku != sku:
                payg_items = get_payg(base_sku)
                ri_items = get_ri(base_sku)

        def _get_price(items, must_be_win=False):
            cands = []
            for i in items:
                prod = i.get("productName", "").lower()
                meter = i.get("meterName", "").lower()
                
                if "low priority" in meter: continue
                if is_spot and "spot" not in meter: continue
                if not is_spot and "spot" in meter: continue
                
                is_win_prod = "windows" in prod
                if must_be_win and not is_win_prod: continue
                if not must_be_win and is_win_prod: continue
                
                if i.get("retailPrice", 0) > 0:
                    cands.append(i["retailPrice"])
            return min(cands) if cands else None

        linux_hr = _get_price(payg_items, must_be_win=False)
        win_hr   = _get_price(payg_items, must_be_win=True)

        api_comp = _hourly_to_monthly(linux_hr) if linux_hr else None
        api_win_tot = _hourly_to_monthly(win_hr) if win_hr else None
        
        result["compute_payg"] = api_comp
        if api_win_tot and api_comp:
            result["windows_license"] = max(0, api_win_tot - api_comp)

        if not is_spot:
            ri1_cands = [i for i in ri_items if i.get("reservationTerm") == "1 Year"]
            ri1_val = _get_price(ri1_cands, must_be_win=False)
            if ri1_val is not None:
                result["compute_ri1"] = ri1_val / 12 

            ri3_cands = [i for i in ri_items if i.get("reservationTerm") == "3 Years"]
            ri3_val = _get_price(ri3_cands, must_be_win=False)
            if ri3_val is not None:
                result["compute_ri3"] = ri3_val / 36

    except Exception as e:
        log.warning(f"VM pricing error for {sku}: {e}")

    return result

def extract_vm_sku(desc):
    if not desc: return None
    m = re.match(r'^\s*[\d,]+\s+([^()]+)\(', desc)
    if m:
        raw = m.group(1).strip()
        norm = re.sub(r'\s+', '_', raw)
        if not norm.lower().startswith("standard_"):
            norm = "Standard_" + norm
        return norm
        
    patterns = [r'^\d+\s+((?:[A-Z][A-Za-z0-9\-]+\s+)+v\d+)', r'^\d+\s+([A-Z][A-Za-z0-9\-]+)\s*\(', r'(Standard_[A-Za-z0-9_\-]+)']
    for pat in patterns:
        m = re.search(pat, desc.strip())
        if m:
            raw = m.group(1).strip()
            norm = re.sub(r'\s+', '_', raw)
            if not norm.startswith("Standard_"): norm = "Standard_" + norm
            return norm
    return None

def extract_quantity(desc):
    if not desc: return 1
    m = re.match(r'^\s*([0-9,]+)\s+', desc)
    if m:
        try:
            return max(1, int(m.group(1).replace(',', '')))
        except ValueError:
            return 1
    return 1

def extract_vcpus(desc):
    if not desc: return None
    m = re.search(r'\((\d+)\s*vCPU', desc, re.IGNORECASE)
    return int(m.group(1)) if m else None

# ── Parsing ────────────────────────────────────────────────────────────────
def parse_format(wb):
    ws = wb.active
    rows = []
    in_data = valid_format = False
    
    def _get_cost(row_data, idx, fallback):
        if len(row_data) > idx and isinstance(row_data[idx], (int, float)) and row_data[idx] > 0:
            return float(row_data[idx])
        return float(fallback)

    for r in ws.iter_rows(values_only=True):
        if not r or r[0] is None: continue
        if str(r[0]).strip().lower() == "service category":
            in_data = valid_format = True
            continue
            
        if not in_data: continue
        svc_cat = str(r[0]).strip() if r[0] else ""
        svc_type = str(r[1]).strip() if r[1] else ""
        region = str(r[3]).strip() if r[3] else ""
        desc = str(r[4]).strip() if r[4] else ""
        cost_raw = r[5]
        
        if svc_cat.lower() in SKIP or region.lower() in SKIP or not isinstance(cost_raw, (int, float)): 
            continue
            
        rows.append({
            "svc_cat": svc_cat, "svc_type": svc_type, "cust_name": str(r[2] or "").strip(),
            "region": region, "desc": desc, "payg": float(cost_raw),
            "ri1": _get_cost(r, 6, cost_raw), 
            "ri3": _get_cost(r, 7, cost_raw), 
            "remarks": "", "sub_rows": []
        })
        
    if not valid_format:
        raise ValueError("Could not find 'Service category' header. Please ensure this is an unmodified Azure Pricing Calculator export.")
    return rows

def classify(rows):
    buckets = {}
    for r in rows:
        key = str(r.get("svc_type") or "").lower().strip()
        sheet = SVC_MAP.get(key, "Others")
        if sheet == "Others":
            for k, v in SVC_MAP.items():
                if k in key: 
                    sheet = v; break
        buckets.setdefault(sheet, []).append(r)
    return buckets

def enrich_vms_concurrent(vm_rows, currency="INR"):
    log.info(f"Querying Azure Retail Pricing API concurrently for {len(vm_rows)} VMs...")
    session = get_http_session()
    cache = {}
    
    def process_row(row):
        desc, region = row["desc"], row["region"]
        os_type, sql_lbl = detect_os(desc), detect_sql_license(desc)
        qty = extract_quantity(desc)
        sku = extract_vm_sku(desc)

        sql_exact, os_exact = get_exact_license_name(desc, os_type)
        row["sql_lbl_exact"] = sql_exact or "SQL License"
        row["os_lbl_exact"]  = os_exact or f"{os_type} License"
        row["api"] = {}

        if not sku:
            is_standalone = False
            if os_type in ["Red Hat", "SUSE"] or sql_lbl: is_standalone = True
            if is_standalone: row["api"] = {"is_standalone": True}
            return row

        is_spot = "spot" in desc.lower()
        p = get_vm_pricing(session, cache, sku, region, is_spot, currency)
        
        api_comp = (p.get("compute_payg") or 0) * qty
        api_win  = (p.get("windows_license") or 0) * qty
        orig_payg = row.get("payg", 0)
        compute_payg = api_comp if api_comp > 0 else orig_payg
        
        unaccounted = orig_payg - compute_payg
        win_lic_payg = 0
        if os_type == "Windows" and api_win > 0:
            win_lic_payg = min(max(0, unaccounted), api_win)
            unaccounted -= win_lic_payg
            
        prem_os_payg = sql_payg = 0
        
        # --- THE SURGICAL FIX ---
        if unaccounted > 5:
            if os_type in ["Red Hat", "SUSE"]:
                prem_os_payg = unaccounted
                unaccounted = 0
            elif sql_exact:
                sql_payg = unaccounted
                unaccounted = 0
            elif os_type == "Linux":
                # Premium OS hidden under "Linux" label by Azure Calculator.
                # Prevent API FX drift from corrupting the OS/Compute split in non-USD currencies.
                vcpus = extract_vcpus(desc)
                if vcpus is not None:
                    os_candidates = [42.05, 29.20] if vcpus <= 4 else [94.90, 73.00]
                else:
                    os_candidates = [42.05, 94.90, 29.20, 73.00]

                # Fetch USD compute in background to reverse-engineer Calculator's exact FX rate
                p_usd = get_vm_pricing(session, cache, sku, region, is_spot, "USD")
                usd_comp = (p_usd.get("compute_payg") or 0) * qty
                
                if usd_comp > 0:
                    api_fx = compute_payg / usd_comp if compute_payg else 1.0
                    best_os = 42.05
                    min_diff = float('inf')
                    
                    for os_usd in os_candidates:
                        calc_fx = orig_payg / (usd_comp + (os_usd * qty))
                        diff = abs(calc_fx - api_fx)
                        if diff < min_diff:
                            min_diff = diff
                            best_os = os_usd
                            
                    calc_fx = orig_payg / (usd_comp + (best_os * qty))
                    exact_os_cost = (best_os * qty) * calc_fx
                    
                    prem_os_payg = min(unaccounted, exact_os_cost)
                else:
                    prem_os_payg = unaccounted
                
                row["os_lbl_exact"] = "Premium OS License" 
                unaccounted -= prem_os_payg
                compute_payg += unaccounted  # FX drift flows cleanly back into Compute
                unaccounted = 0
            else:
                compute_payg += unaccounted 
                unaccounted = 0
        # ------------------------
                
        p["compute_payg_final"] = compute_payg
        p["win_lic_payg_final"] = win_lic_payg
        p["sql_payg_final"]     = sql_payg
        p["prem_os_payg_final"] = prem_os_payg
        
        if p.get("compute_ri1"): p["compute_ri1"] *= qty
        if p.get("compute_ri3"): p["compute_ri3"] *= qty
        
        row["api"] = p
        return row

    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = {executor.submit(process_row, r): r for r in vm_rows}
        for future in as_completed(futures):
            pass 

    return vm_rows

# ── Output ────────────────────────────────────────────────────────────────
def write_res_header(ws):
    ws.merge_cells("F1:H1")
    hdr(ws["F1"], "Monthly Cost")
    for addr in ["G1","H1"]: ws[addr].border=BORDER
    for ci, h in enumerate(["Service category","Service type","Custom name", "Region","Description","PAYG", "1 Year RI Model","3 Year RI Model","Remarks"], 1):
        hdr(ws.cell(2, ci), h, wrap=True)
    ws.row_dimensions[2].height = 28.8

def write_vm_sheet(wb, rows):
    ws = wb.create_sheet("Virtual Machines")
    write_res_header(ws)
    widths(ws, {"A":15, "B":15, "C":14, "D":12, "E":55, "F":13, "G":14, "H":14, "I":42})

    ri, total_payg, total_ri1, total_ri3 = 3, 0.0, 0.0, 0.0

    for row in rows:
        p = row.get("api", {})
        if p.get("is_standalone"):
            payg = row.get("payg", 0)
            ri1 = row.get("ri1", payg)
            ri3 = row.get("ri3", payg)
            
            vals = [row["svc_cat"], row["svc_type"], row["cust_name"], row["region"], row["desc"],
                    round(payg,2), round(ri1,2), round(ri3,2), row.get("remarks","")]
            for ci, v in enumerate(vals, 1): dat(ws.cell(ri, ci), v, align="right" if ci>=6 and isinstance(v,float) else "left")
            ri += 1
            total_payg += payg; total_ri1 += ri1; total_ri3 += ri3
            continue

        compute_payg = p.get("compute_payg_final", row.get("payg",0))
        win_lic_payg = p.get("win_lic_payg_final", 0)
        prem_os_payg = p.get("prem_os_payg_final", 0)
        
        sql_rows_b = [s.copy() for s in row.get("sub_rows",[]) if "sql" in s["desc"].lower()]
        sql_payg_deduced = p.get("sql_payg_final", 0)
        
        if not sql_rows_b and sql_payg_deduced > 0:
            sql_rows_b.append({
                "desc": row.get("sql_lbl_exact", "SQL License"),
                "payg": sql_payg_deduced, "ri1": sql_payg_deduced, "ri3": sql_payg_deduced, "is_api": True
            })

        sql_payg_total = sum(s["payg"] for s in sql_rows_b if s.get("payg"))
        sql_ri1_total  = sum(s.get("ri1") or s["payg"] for s in sql_rows_b)
        sql_ri3_total  = sum(s.get("ri3") or s["payg"] for s in sql_rows_b)

        api_ri1 = p.get("compute_ri1")
        compute_ri1 = api_ri1 if api_ri1 is not None else (row.get("ri1") if row.get("ri1") is not None else compute_payg)
        
        api_ri3 = p.get("compute_ri3")
        compute_ri3 = api_ri3 if api_ri3 is not None else (row.get("ri3") if row.get("ri3") is not None else compute_payg)
        
        vm_payg = compute_payg + win_lic_payg + prem_os_payg + sql_payg_total
        vm_ri1  = compute_ri1  + win_lic_payg + prem_os_payg + sql_ri1_total
        vm_ri3  = compute_ri3  + win_lic_payg + prem_os_payg + sql_ri3_total

        vals = [row["svc_cat"], row["svc_type"], row["cust_name"], row["region"], row["desc"],
                round(compute_payg,2), round(compute_ri1,2), round(compute_ri3,2), row.get("remarks","")]
        for ci, v in enumerate(vals, 1): dat(ws.cell(ri, ci), v, align="right" if ci>=6 and isinstance(v,float) else "left")
        ri += 1

        if win_lic_payg > 0:
            sub = ["","","","","Windows License", round(win_lic_payg,2), round(win_lic_payg,2), round(win_lic_payg,2), "License Cost"]
            for ci, v in enumerate(sub, 1): dat(ws.cell(ri, ci), v, italic=True, color="595959", align="right" if ci>=6 and isinstance(v,float) else "left")
            ri += 1

        if prem_os_payg > 0:
            sub = ["","","","", row.get("os_lbl_exact", "Premium OS License"), round(prem_os_payg,2), round(prem_os_payg,2), round(prem_os_payg,2), "License Cost"]
            for ci, v in enumerate(sub, 1): dat(ws.cell(ri, ci), v, italic=True, color="595959", align="right" if ci>=6 and isinstance(v,float) else "left")
            ri += 1

        for s in sql_rows_b:
            rmk = "License Cost" if s.get("is_api") else s.get("remarks", "")
            sub = ["","","","", s["desc"], round(s["payg"],2) if s.get("payg") else None, round(s.get("ri1") or s["payg"],2) if s.get("payg") else None, round(s.get("ri3") or s["payg"],2) if s.get("payg") else None, rmk]
            for ci, v in enumerate(sub, 1): dat(ws.cell(ri, ci), v, italic=True, color="595959", align="right" if ci>=6 and isinstance(v,float) else "left")
            ri += 1

        for s in [s for s in row.get("sub_rows",[]) if "sql" not in s["desc"].lower()]:
            sub = ["","","","", s["desc"], round(s["payg"],2) if s.get("payg") else None, round(s.get("ri1") or s["payg"],2) if s.get("payg") else None, round(s.get("ri3") or s["payg"],2) if s.get("payg") else None, s.get("remarks","")]
            for ci, v in enumerate(sub, 1): dat(ws.cell(ri, ci), v, italic=True, color="595959", align="right" if ci>=6 and isinstance(v,float) else "left")
            ri += 1

        total_payg += vm_payg; total_ri1 += vm_ri1; total_ri3 += vm_ri3

    ws.cell(ri, 5, "Total").font = _f(bold=True)
    ws.cell(ri, 5).border = BORDER
    for ci, v in [(6,total_payg), (7,total_ri1), (8,total_ri3)]: tot(ws.cell(ri, ci), round(v,2))

    return total_payg, total_ri1, total_ri3

def write_generic_sheet(wb, sheet_name, rows):
    ws = wb.create_sheet(sheet_name)
    write_res_header(ws)
    widths(ws, {"A":15, "B":14, "C":22, "D":12, "E":60, "F":13, "G":14, "H":14, "I":40})

    ri, tp, tr1, tr3 = 3, 0.0, 0.0, 0.0
    for row in rows:
        payg = row.get("payg", 0)
        ri1, ri3 = row.get("ri1", payg), row.get("ri3", payg)
        
        vals = [row["svc_cat"], row["svc_type"], row["cust_name"], row["region"], row["desc"], round(payg,2), round(ri1,2), round(ri3,2), row.get("remarks","")]
        for ci, v in enumerate(vals, 1): dat(ws.cell(ri, ci), v, align="right" if ci>=6 and isinstance(v,float) else "left")
        
        tp += payg; tr1 += ri1; tr3 += ri3; ri += 1

    ws.cell(ri, 5, "Total").font = _f(bold=True)
    ws.cell(ri, 5).border = BORDER
    for ci, v in [(6,tp), (7,tr1), (8,tr3)]: tot(ws.cell(ri, ci), round(v,2))
    return tp, tr1, tr3

def convert(input_path, output_path, currency="INR"):
    wb_in = load_workbook(input_path, data_only=True)
    rows = parse_format(wb_in)
    buckets = classify(rows)
    
    if "Virtual Machines" in buckets:
        enrich_vms_concurrent(buckets["Virtual Machines"], currency)

    wb_out = Workbook()
    wb_out.remove(wb_out.active)
    totals = {}

    for sname in SHEET_ORDER:
        if sname not in buckets: continue
        if sname == "Virtual Machines":
            totals[sname] = write_vm_sheet(wb_out, buckets[sname])
        else:
            totals[sname] = write_generic_sheet(wb_out, sname, buckets[sname])

    ws = wb_out.create_sheet("Summary", 0)
    ws.merge_cells("A1:A2"); ws.merge_cells("B1:B2"); ws.merge_cells("C1:E1"); ws.merge_cells("F1:F2")
    for addr, val, al in [("A1","Sl No","left"),("B1","Service Name","left"), ("C1","Monthly Cost","center"),("F1","Remarks","left")]:
        c=ws[addr]; c.value=val; c.font=_f(bold=True); c.alignment=_al(al,"center"); c.border=BORDER
    for addr, val in [("C2","PAYG"),("D2","1 YR RI Model"),("E2","3 YR RI Model")]:
        c=ws[addr]; c.value=val; c.font=_f(bold=True); c.alignment=_al("center","center"); c.border=BORDER

    gp=gr1=gr3=0.0; ri=3
    for sl, (sname, (payg, ri1, ri3)) in enumerate(totals.items(), 1):
        c=ws.cell(ri,1,sl); c.border=BORDER; c.alignment=_al("center")
        c=ws.cell(ri,2,sname); c.border=BORDER
        for ci, v in [(3,payg),(4,ri1),(5,ri3)]:
            c=ws.cell(ri,ci,round(v,2)); c.alignment=_al("right"); c.border=BORDER; c.number_format=NUM
        ws.cell(ri,6).border=BORDER
        gp+=payg; gr1+=ri1; gr3+=ri3; ri+=1

    for ci, v in [(3,gp),(4,gr1),(5,gr3)]: tot(ws.cell(ri, ci), round(v,2))
    widths(ws, {"A":5.5, "B":22, "C":14, "D":14, "E":14, "F":48})
    wb_out.save(output_path)

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python convert.py input.xlsx [output.xlsx]")
        sys.exit(1)
    convert(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else "output.xlsx")
